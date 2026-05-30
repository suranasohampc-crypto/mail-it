import streamlit as st
import smtplib
import time
import io
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Mail It", page_icon="📧")

# ── PASSWORD PROTECTION ────────────────────────────────────────
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("📧 Mail It")
    st.subheader("🔒 Login to Continue")
    pwd = st.text_input("Enter App Password", type="password")
    if st.button("Login", use_container_width=True, type="primary"):
        if pwd == st.secrets["APP_PASSWORD"]:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("❌ Incorrect password. Please try again.")
    st.stop()

# ── MAIN APP ───────────────────────────────────────────────────
st.title("📧 Mail It")
st.caption("Upload Excel → Send Emails → Download Updated File")

if st.button("🚪 Logout", help="Click to logout"):
    st.session_state.authenticated = False
    st.rerun()

# ── SIDEBAR ────────────────────────────────────────────────────
with st.sidebar:
    st.header("🔐 Gmail Settings")
    gmail_id = st.text_input("Gmail Address", placeholder="you@gmail.com")
    app_password = st.text_input("App Password", type="password", placeholder="abcd efgh ijkl mnop")
    delay = st.slider("Delay between emails (sec)", 1, 10, 2)

    st.header("👤 Your Details")
    your_name = st.text_input("Your Name", placeholder="e.g. Rajesh Kumar")
    your_firm = st.text_input("Your Firm Name", placeholder="e.g. Kumar & Associates")
    your_phone = st.text_input("Your Phone", placeholder="e.g. 98765 43210")

    st.header("✉️ Email Content")
    subject = st.text_input("Subject", value="Request for Income Tax Return Data - FY 2024-25")
    signature = f"{your_name}\n{your_firm}\n{your_phone}" if any([your_name, your_firm, your_phone]) else "[Your Name]\n[Your Firm]\n[Your Phone]"
    body = st.text_area("Email Body", height=250, value=f"""Dear Client,

We hope this message finds you well.

As the income tax filing deadline approaches, we kindly request you to share your ITR data for FY 2024-25 at the earliest.

Please send the following documents:
- Form 16 / Salary slips
- Bank statements
- Investment proofs (80C, 80D, etc.)
- Any other income details

Regards,
{signature}""")

# ── MAIN AREA ──────────────────────────────────────────────────
st.subheader("📂 Upload Excel File")
st.caption("Excel must have columns: Name, Email. For different attachments per client, add an 'Attachment' column with file names.")
uploaded_file = st.file_uploader("Choose your Excel file", type=["xlsx"])

# ── ATTACHMENT SECTION ─────────────────────────────────────────
st.subheader("📎 Attachments")
attachment_mode = st.radio(
    "Attachment Mode",
    ["Same attachment for all clients", "Different attachment per client (from Excel)"],
    horizontal=True
)

common_attachment = None
client_attachments = {}

if attachment_mode == "Same attachment for all clients":
    common_attachment = st.file_uploader(
        "Upload attachment (optional)",
        type=["pdf", "xlsx", "xls", "docx", "doc", "png", "jpg", "jpeg"],
        key="common"
    )
    if common_attachment:
        st.success(f"✅ Attachment ready: {common_attachment.name}")

else:
    st.info("📋 Upload all client files below. Make sure filenames match the 'Attachment' column in your Excel.")
    uploaded_attachments = st.file_uploader(
        "Upload client attachments",
        type=["pdf", "xlsx", "xls", "docx", "doc", "png", "jpg", "jpeg"],
        accept_multiple_files=True,
        key="multiple"
    )
    if uploaded_attachments:
        for f in uploaded_attachments:
            client_attachments[f.name] = f
        st.success(f"✅ {len(client_attachments)} files uploaded: {', '.join(client_attachments.keys())}")

# ── PROCESS EXCEL ──────────────────────────────────────────────
if uploaded_file:
    wb = load_workbook(uploaded_file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    try:
        name_col = next(i for i, h in enumerate(headers) if h and "name" in str(h).lower())
        email_col = next(i for i, h in enumerate(headers) if h and "email" in str(h).lower())
    except StopIteration:
        st.error("❌ Could not find 'Name' or 'Email' columns. Please check column headers.")
        st.stop()

    # Check for attachment column
    attach_col = next((i for i, h in enumerate(headers) if h and "attachment" in str(h).lower()), None)

    if "Status" not in headers:
        ws.cell(row=1, column=len(headers) + 1).value = "Status"
        status_col = len(headers) + 1
    else:
        status_col = headers.index("Status") + 1

    data_preview = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[email_col]:
            preview_row = {"Name": row[name_col], "Email": row[email_col]}
            if attach_col is not None:
                preview_row["Attachment"] = row[attach_col]
            data_preview.append(preview_row)

    st.success(f"✅ Found **{len(data_preview)} clients** in the file.")
    st.dataframe(data_preview, use_container_width=True)

    if st.button("🚀 Send Emails", use_container_width=True, type="primary"):
        if not gmail_id or not app_password:
            st.error("Please fill in Gmail Address and App Password in the sidebar.")
            st.stop()

        def send_email(to, attachment_file=None, attachment_name=None):
            msg = MIMEMultipart()
            msg['Subject'] = subject
            msg['From'] = gmail_id
            msg['To'] = to
            msg.attach(MIMEText(body, 'plain'))

            if attachment_file is not None:
                attachment_file.seek(0)
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment_file.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{attachment_name}"')
                msg.attach(part)

            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
                s.login(gmail_id, app_password)
                s.send_message(msg)

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        progress = st.progress(0)
        log_box = st.empty()
        sent = failed = 0
        rows = list(ws.iter_rows(min_row=2))
        total = len([r for r in rows if r[email_col].value])

        for i, row in enumerate(rows):
            email = row[email_col].value
            if not email:
                continue

            # Determine attachment for this client
            attachment_file = None
            attachment_name = None

            if attachment_mode == "Same attachment for all clients" and common_attachment:
                attachment_file = common_attachment
                attachment_name = common_attachment.name
            elif attachment_mode == "Different attachment per client (from Excel)" and attach_col is not None:
                client_file_name = row[attach_col].value if row[attach_col].value else None
                if client_file_name and client_file_name in client_attachments:
                    attachment_file = client_attachments[client_file_name]
                    attachment_name = client_file_name

            try:
                send_email(str(email).strip(), attachment_file, attachment_name)
                status = "Sent ✅" + (f" (with {attachment_name})" if attachment_name else "")
                ws.cell(row=row[0].row, column=status_col).value = status
                ws.cell(row=row[0].row, column=status_col).fill = green_fill
                sent += 1
                log_box.success(f"[{sent + failed}/{total}] ✅ Sent to {email}" + (f" + 📎 {attachment_name}" if attachment_name else ""))
            except Exception as e:
                ws.cell(row=row[0].row, column=status_col).value = "Failed ❌"
                ws.cell(row=row[0].row, column=status_col).fill = red_fill
                failed += 1
                log_box.error(f"[{sent + failed}/{total}] ❌ Failed: {email} — {e}")

            progress.progress((sent + failed) / total)
            time.sleep(delay)

        st.balloons()
        st.success(f"✅ Done! Sent: {sent} | Failed: {failed}")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.download_button(
            label="📥 Download Updated Excel File",
            data=output,
            file_name="email_status_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
